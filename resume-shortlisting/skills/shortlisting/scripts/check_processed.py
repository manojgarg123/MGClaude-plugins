#!/usr/bin/env python3
"""
check_processed.py — List resume filenames already recorded in the candidate tracker.

Usage:
  python check_processed.py --tracker "<path-to-Candidate_Tracker.xlsx>"

Prints one filename per line (exactly as stored in the "Resume File" column).
Exits cleanly with no output if the tracker doesn't exist yet.
"""

import argparse
import os
import sys


def main():
    parser = argparse.ArgumentParser(
        description="List already-processed resume filenames from the tracker")
    parser.add_argument("--tracker", required=True, help="Path to Candidate_Tracker.xlsx")
    args = parser.parse_args()

    if not os.path.exists(args.tracker):
        # No tracker yet — nothing has been processed
        sys.exit(0)

    try:
        import openpyxl
    except ImportError:
        os.system("pip install openpyxl --break-system-packages -q")
        import openpyxl

    try:
        wb = openpyxl.load_workbook(args.tracker, data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        print(f"WARNING: Could not read tracker: {e}", file=sys.stderr)
        sys.exit(0)

    headers = [cell.value for cell in ws[1]]

    try:
        resume_col_idx = headers.index("Resume File")
    except ValueError:
        # Column not found — treat as empty tracker
        sys.exit(0)

    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[resume_col_idx] if resume_col_idx < len(row) else None
        if val and str(val).strip():
            print(str(val).strip())


if __name__ == "__main__":
    main()
