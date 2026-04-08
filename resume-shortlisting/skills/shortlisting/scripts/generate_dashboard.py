#!/usr/bin/env python3
"""
generate_dashboard.py — Generate a standalone HTML analytics dashboard from the candidate tracker.

Usage:
  python generate_dashboard.py \
    --tracker "<path-to-Candidate_Tracker.xlsx>" \
    --output  "<path-to-output.html>"

The output is a single self-contained HTML file (Chart.js loaded from CDN) that can be
opened directly in any browser with no server required.
"""

import argparse
import json
import os
import re
from collections import Counter
from datetime import datetime


# ─────────────────────────────────────────────
# 1. Data extraction
# ─────────────────────────────────────────────

def load_tracker(tracker_path):
    try:
        import openpyxl
    except ImportError:
        os.system("pip install openpyxl --break-system-packages -q")
        import openpyxl

    wb = openpyxl.load_workbook(tracker_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    def col(name):
        """Return 0-based index for a column by header name."""
        try:
            return headers.index(name)
        except ValueError:
            return None

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None for cell in row):
            continue
        rows.append(row)

    return headers, rows, col


def parse_years(val):
    """Try to extract a numeric year value from a cell."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        m = re.search(r"(\d+(?:\.\d+)?)", str(val))
        return float(m.group(1)) if m else None


def build_data(tracker_path):
    headers, rows, col = load_tracker(tracker_path)

    # Column indices
    i_pos_title   = col("Position Title")
    i_rank        = col("Rank")
    i_name        = col("Candidate Name")
    i_org         = col("Current Organisation")
    i_qual        = col("Highest Qualification")
    i_yoe         = col("Years of Experience")
    i_skills      = col("Skills Score (/40)")
    i_exp         = col("Experience Score (/30)")
    i_edu         = col("Education Score (/20)")
    i_fit         = col("Overall Fit Score (/10)")
    i_total       = col("Total Score (/100)")
    i_strengths   = col("Key Strengths")
    i_gaps        = col("Key Gaps")
    i_date        = col("Date Added")

    def g(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    # Position info from first data row
    first = rows[0] if rows else [None] * len(headers)
    position = {
        "title": g(first, i_pos_title) or "N/A",
    }

    candidates = []
    for row in rows:
        total = g(row, i_total)
        try:
            total = int(total)
        except (TypeError, ValueError):
            total = 0

        candidates.append({
            "rank":      g(row, i_rank),
            "name":      g(row, i_name)      or "Unknown",
            "org":       g(row, i_org)       or "N/A",
            "qual":      g(row, i_qual)       or "N/A",
            "yoe":       parse_years(g(row, i_yoe)),
            "skills":    g(row, i_skills)    or 0,
            "exp":       g(row, i_exp)       or 0,
            "edu":       g(row, i_edu)       or 0,
            "fit":       g(row, i_fit)       or 0,
            "total":     total,
            "strengths": g(row, i_strengths) or "",
            "gaps":      g(row, i_gaps)      or "",
            "date":      str(g(row, i_date)) if g(row, i_date) else "",
        })

    # Sort by total score desc
    candidates.sort(key=lambda c: c["total"], reverse=True)

    # ── Derived stats ──────────────────────────────────────
    total_count    = len(candidates)
    shortlisted    = [c for c in candidates if c["total"] >= 75]
    on_hold        = [c for c in candidates if 50 <= c["total"] < 75]
    not_suitable   = [c for c in candidates if c["total"] < 50]

    # Average scores
    def avg(lst, key):
        vals = [c[key] for c in lst if c[key] is not None]
        return round(sum(vals) / len(vals), 1) if vals else 0

    avg_skills = avg(candidates, "skills")
    avg_exp    = avg(candidates, "exp")
    avg_edu    = avg(candidates, "edu")
    avg_fit    = avg(candidates, "fit")

    # Experience distribution
    bands = {"0–2 yrs": 0, "3–5 yrs": 0, "6–10 yrs": 0, "10+ yrs": 0}
    for c in candidates:
        y = c["yoe"]
        if y is None:
            continue
        if y <= 2:
            bands["0–2 yrs"] += 1
        elif y <= 5:
            bands["3–5 yrs"] += 1
        elif y <= 10:
            bands["6–10 yrs"] += 1
        else:
            bands["10+ yrs"] += 1

    # Common gaps
    all_gaps = []
    for c in candidates:
        for g_phrase in c["gaps"].split(","):
            g_phrase = g_phrase.strip()
            if g_phrase and g_phrase.lower() not in ("n/a", "none", ""):
                all_gaps.append(g_phrase.lower())
    gap_counts = Counter(all_gaps).most_common(10)

    top5 = candidates[:5]

    generated_at = datetime.now().strftime("%d %b %Y, %H:%M")

    return {
        "position":      position,
        "total":         total_count,
        "shortlisted":   len(shortlisted),
        "on_hold":       len(on_hold),
        "not_suitable":  len(not_suitable),
        "avg_skills":    avg_skills,
        "avg_exp":       avg_exp,
        "avg_edu":       avg_edu,
        "avg_fit":       avg_fit,
        "exp_bands":     bands,
        "gap_counts":    gap_counts,
        "top5":          top5,
        "candidates":    candidates,
        "generated_at":  generated_at,
    }


# ─────────────────────────────────────────────
# 2. HTML generation
# ─────────────────────────────────────────────

def score_color(score):
    if score >= 75:
        return "#00B050"
    elif score >= 50:
        return "#FF9900"
    return "#FF0000"


def score_badge_class(score):
    if score >= 75:
        return "badge-green"
    elif score >= 50:
        return "badge-amber"
    return "badge-red"


def render_top5_cards(top5):
    cards = []
    for c in top5:
        sc = score_color(c["total"])
        badge = score_badge_class(c["total"])
        strengths_html = ""
        if c["strengths"]:
            tags = [f'<span class="tag">{s.strip()}</span>'
                    for s in c["strengths"].split(",") if s.strip()]
            strengths_html = '<div class="tags">' + "".join(tags) + "</div>"

        yoe_str = f"{int(c['yoe'])} yrs" if c["yoe"] is not None else "N/A"

        cards.append(f"""
        <div class="cand-card">
          <div class="cand-header">
            <div>
              <div class="cand-name">{c['name']}</div>
              <div class="cand-meta">{c['org']} &nbsp;|&nbsp; {c['qual']} &nbsp;|&nbsp; {yoe_str}</div>
            </div>
            <div class="score-badge {badge}">{c['total']}<span class="score-denom">/100</span></div>
          </div>
          <div class="score-bar-row">
            <span class="score-label">Skills</span>
            <div class="score-bar"><div class="score-fill" style="width:{int(c['skills']/40*100)}%;background:#2563eb"></div></div>
            <span class="score-val">{c['skills']}/40</span>
            <span class="score-label">Exp</span>
            <div class="score-bar"><div class="score-fill" style="width:{int(c['exp']/30*100)}%;background:#7c3aed"></div></div>
            <span class="score-val">{c['exp']}/30</span>
            <span class="score-label">Edu</span>
            <div class="score-bar"><div class="score-fill" style="width:{int(c['edu']/20*100)}%;background:#0891b2"></div></div>
            <span class="score-val">{c['edu']}/20</span>
            <span class="score-label">Fit</span>
            <div class="score-bar"><div class="score-fill" style="width:{int(c['fit']/10*100)}%;background:#059669"></div></div>
            <span class="score-val">{c['fit']}/10</span>
          </div>
          {strengths_html}
        </div>""")
    return "\n".join(cards)


def render_gap_list(gap_counts):
    if not gap_counts:
        return "<p style='color:#888'>No gaps data available.</p>"
    max_count = gap_counts[0][1] if gap_counts else 1
    items = []
    for gap, count in gap_counts:
        pct = int(count / max_count * 100)
        items.append(f"""
        <div class="gap-item">
          <div class="gap-label">{gap.title()}</div>
          <div class="gap-bar-wrap">
            <div class="gap-bar-fill" style="width:{pct}%"></div>
          </div>
          <div class="gap-count">{count} candidate{"s" if count != 1 else ""}</div>
        </div>""")
    return "\n".join(items)


def render_all_candidates_table(candidates):
    rows = []
    for c in candidates:
        sc = score_color(c["total"])
        badge = score_badge_class(c["total"])
        yoe_str = f"{int(c['yoe'])}" if c["yoe"] is not None else "N/A"
        rows.append(f"""
        <tr>
          <td>{c.get('rank', '')}</td>
          <td><strong>{c['name']}</strong></td>
          <td>{c['org']}</td>
          <td>{c['qual']}</td>
          <td>{yoe_str}</td>
          <td>{c['skills']}/40</td>
          <td>{c['exp']}/30</td>
          <td>{c['edu']}/20</td>
          <td>{c['fit']}/10</td>
          <td><span class="tbl-badge {badge}">{c['total']}</span></td>
        </tr>""")
    return "\n".join(rows)


def generate_html(data):
    pos          = data["position"]
    exp_labels   = json.dumps(list(data["exp_bands"].keys()))
    exp_values   = json.dumps(list(data["exp_bands"].values()))
    top5_cards   = render_top5_cards(data["top5"])
    gap_list     = render_gap_list(data["gap_counts"])
    all_cands    = render_all_candidates_table(data["candidates"])

    shortlist_pct = round(data["shortlisted"] / data["total"] * 100) if data["total"] else 0
    on_hold_pct   = round(data["on_hold"]     / data["total"] * 100) if data["total"] else 0
    unsuitable_pct = round(data["not_suitable"]/ data["total"] * 100) if data["total"] else 0

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>Shortlisting Dashboard – {pos['title']}</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #f0f4f8; color: #1e293b; }}
  a {{ color: inherit; text-decoration: none; }}

  /* ── Top bar ── */
  .topbar {{ background: linear-gradient(135deg, #1e3a5f 0%, #2563eb 100%);
             color: #fff; padding: 24px 40px; }}
  .topbar h1 {{ font-size: 1.7rem; font-weight: 700; margin-bottom: 4px; }}
  .topbar .meta {{ font-size: 0.85rem; opacity: .75; }}

  /* pill row removed — position shown in topbar only */

  /* ── Page content ── */
  .page {{ max-width: 1300px; margin: 0 auto; padding: 28px 24px; }}

  /* ── Stat cards ── */
  .stat-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
               gap: 16px; margin-bottom: 28px; }}
  .stat-card {{ background: #fff; border-radius: 14px; padding: 22px 20px;
               box-shadow: 0 1px 4px rgba(0,0,0,.07); text-align: center; }}
  .stat-card .num {{ font-size: 2.2rem; font-weight: 800; line-height: 1; }}
  .stat-card .lbl {{ font-size: 0.8rem; color: #64748b; margin-top: 6px; text-transform: uppercase; letter-spacing: .04em; }}
  .num-total {{ color: #1e3a5f; }}
  .num-green {{ color: #00B050; }}
  .num-amber {{ color: #FF9900; }}
  .num-red   {{ color: #FF0000; }}

  /* ── Section titles ── */
  .section-title {{ font-size: 1.1rem; font-weight: 700; color: #1e3a5f;
                    margin-bottom: 16px; padding-bottom: 8px;
                    border-bottom: 2px solid #e2e8f0; }}

  /* ── Two-col layout ── */
  .two-col {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 28px; }}
  @media (max-width: 860px) {{ .two-col {{ grid-template-columns: 1fr; }} }}

  .card {{ background: #fff; border-radius: 14px; padding: 22px;
           box-shadow: 0 1px 4px rgba(0,0,0,.07); }}

  /* ── Chart containers ── */
  .chart-wrap {{ position: relative; height: 240px; }}

  /* ── Top 5 cards ── */
  .cand-card {{ background: #fff; border-radius: 14px; padding: 18px 20px;
               box-shadow: 0 1px 4px rgba(0,0,0,.07); margin-bottom: 14px; }}
  .cand-header {{ display: flex; justify-content: space-between; align-items: flex-start;
                  margin-bottom: 12px; }}
  .cand-name {{ font-size: 1.05rem; font-weight: 700; color: #1e3a5f; }}
  .cand-meta {{ font-size: 0.8rem; color: #64748b; margin-top: 3px; }}

  .score-badge {{ font-size: 1.6rem; font-weight: 800; line-height: 1; text-align: right; }}
  .score-denom {{ font-size: 0.75rem; font-weight: 400; color: #94a3b8; }}
  .badge-green {{ color: #00B050; }}
  .badge-amber {{ color: #FF9900; }}
  .badge-red   {{ color: #FF0000; }}

  .score-bar-row {{ display: flex; align-items: center; gap: 8px; flex-wrap: wrap; margin-bottom: 10px; }}
  .score-label {{ font-size: 0.72rem; color: #64748b; min-width: 28px; }}
  .score-bar {{ flex: 1; min-width: 60px; height: 6px; background: #e2e8f0; border-radius: 999px; overflow: hidden; }}
  .score-fill {{ height: 100%; border-radius: 999px; }}
  .score-val {{ font-size: 0.72rem; color: #475569; min-width: 32px; text-align: right; }}

  .tags {{ display: flex; flex-wrap: wrap; gap: 6px; }}
  .tag {{ background: #eff6ff; color: #1d4ed8; border-radius: 999px;
          padding: 3px 10px; font-size: 0.75rem; }}

  /* ── Gap list ── */
  .gap-item {{ display: flex; align-items: center; gap: 12px; margin-bottom: 10px; }}
  .gap-label {{ min-width: 180px; font-size: 0.85rem; color: #334155; }}
  .gap-bar-wrap {{ flex: 1; height: 8px; background: #e2e8f0; border-radius: 999px; overflow: hidden; }}
  .gap-bar-fill {{ height: 100%; background: linear-gradient(90deg, #f43f5e, #fb923c); border-radius: 999px; }}
  .gap-count {{ font-size: 0.78rem; color: #64748b; min-width: 80px; text-align: right; }}

  /* ── All candidates table ── */
  .table-wrap {{ overflow-x: auto; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 0.82rem; }}
  th {{ background: #1e3a5f; color: #fff; padding: 10px 12px; text-align: left;
        font-weight: 600; white-space: nowrap; }}
  td {{ padding: 9px 12px; border-bottom: 1px solid #e2e8f0; vertical-align: top; }}
  tr:nth-child(even) td {{ background: #f8fafc; }}
  .tbl-badge {{ font-weight: 700; font-size: 0.88rem; }}
  .tbl-badge.badge-green {{ color: #00B050; }}
  .tbl-badge.badge-amber {{ color: #FF9900; }}
  .tbl-badge.badge-red   {{ color: #FF0000; }}

  /* ── Footer ── */
  .footer {{ text-align: center; color: #94a3b8; font-size: 0.78rem; padding: 24px; }}
</style>
</head>
<body>

<!-- Top bar -->
<div class="topbar">
  <h1>Shortlisting Dashboard &mdash; {pos['title']}</h1>
  <div class="meta">Generated {data['generated_at']}</div>
</div>

<div class="page">

  <!-- Summary stats -->
  <div class="stat-grid" style="margin-bottom:28px">
    <div class="stat-card">
      <div class="num num-total">{data['total']}</div>
      <div class="lbl">Total Resumes</div>
    </div>
    <div class="stat-card">
      <div class="num num-green">{data['shortlisted']}</div>
      <div class="lbl">Shortlisted (&ge;75)</div>
    </div>
    <div class="stat-card">
      <div class="num num-amber">{data['on_hold']}</div>
      <div class="lbl">On Hold (50–74)</div>
    </div>
    <div class="stat-card">
      <div class="num num-red">{data['not_suitable']}</div>
      <div class="lbl">Not Suitable (&lt;50)</div>
    </div>
    <div class="stat-card">
      <div class="num" style="color:#2563eb">{shortlist_pct}%</div>
      <div class="lbl">Shortlist Rate</div>
    </div>
  </div>

  <!-- Charts row -->
  <div class="two-col">
    <div class="card">
      <div class="section-title">Candidate Distribution</div>
      <div class="chart-wrap">
        <canvas id="donutChart"></canvas>
      </div>
    </div>
    <div class="card">
      <div class="section-title">Avg Score by Dimension</div>
      <div class="chart-wrap">
        <canvas id="scoreChart"></canvas>
      </div>
    </div>
  </div>

  <!-- Experience distribution -->
  <div class="card" style="margin-bottom:28px">
    <div class="section-title">Experience Distribution</div>
    <div class="chart-wrap" style="height:200px">
      <canvas id="expChart"></canvas>
    </div>
  </div>

  <!-- Top 5 candidates -->
  <div class="section-title">Top 5 Candidates</div>
  {top5_cards}

  <!-- Common gaps -->
  <div class="card" style="margin:28px 0">
    <div class="section-title">Most Common Gaps Across Candidates</div>
    {gap_list}
  </div>

  <!-- All candidates table -->
  <div class="card" style="margin-bottom:28px">
    <div class="section-title">All Candidates — Ranked</div>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th>#</th><th>Name</th><th>Current Org</th><th>Qualification</th>
            <th>Exp (yrs)</th><th>Skills</th><th>Exp</th><th>Edu</th><th>Fit</th><th>Total</th>
          </tr>
        </thead>
        <tbody>
          {all_cands}
        </tbody>
      </table>
    </div>
  </div>

</div>

<div class="footer">Generated by the Shortlisting Skill &bull; {data['generated_at']}</div>

<script>
// Donut chart — candidate distribution
new Chart(document.getElementById('donutChart'), {{
  type: 'doughnut',
  data: {{
    labels: ['Shortlisted (≥75)', 'On Hold (50–74)', 'Not Suitable (<50)'],
    datasets: [{{
      data: [{data['shortlisted']}, {data['on_hold']}, {data['not_suitable']}],
      backgroundColor: ['#00B050', '#FF9900', '#FF0000'],
      borderWidth: 2,
      borderColor: '#fff'
    }}]
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    plugins: {{
      legend: {{ position: 'bottom', labels: {{ font: {{ size: 12 }} }} }}
    }}
  }}
}});

// Bar chart — average scores by dimension
new Chart(document.getElementById('scoreChart'), {{
  type: 'bar',
  data: {{
    labels: ['Skills (/40)', 'Experience (/30)', 'Education (/20)', 'Fit (/10)'],
    datasets: [{{
      label: 'Avg score',
      data: [{data['avg_skills']}, {data['avg_exp']}, {data['avg_edu']}, {data['avg_fit']}],
      backgroundColor: ['#2563eb', '#7c3aed', '#0891b2', '#059669'],
      borderRadius: 6
    }},
    {{
      label: 'Max possible',
      data: [40, 30, 20, 10],
      backgroundColor: ['rgba(37,99,235,.15)', 'rgba(124,58,237,.15)',
                        'rgba(8,145,178,.15)', 'rgba(5,150,105,.15)'],
      borderRadius: 6
    }}]
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    plugins: {{ legend: {{ position: 'bottom', labels: {{ font: {{ size: 11 }} }} }} }},
    scales: {{ y: {{ beginAtZero: true }} }}
  }}
}});

// Bar chart — experience distribution
new Chart(document.getElementById('expChart'), {{
  type: 'bar',
  data: {{
    labels: {exp_labels},
    datasets: [{{
      label: 'Candidates',
      data: {exp_values},
      backgroundColor: '#6366f1',
      borderRadius: 6
    }}]
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    plugins: {{ legend: {{ display: false }} }},
    scales: {{ y: {{ beginAtZero: true, ticks: {{ stepSize: 1 }} }} }}
  }}
}});
</script>
</body>
</html>
"""


# ─────────────────────────────────────────────
# 3. Entry point
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate HTML dashboard from candidate tracker")
    parser.add_argument("--tracker", required=True, help="Path to Candidate_Tracker.xlsx")
    parser.add_argument("--output",  required=True, help="Path for output .html file")
    args = parser.parse_args()

    if not os.path.exists(args.tracker):
        print(f"ERROR: Tracker not found: {args.tracker}", flush=True)
        raise SystemExit(1)

    print("Reading tracker…", flush=True)
    data = build_data(args.tracker)

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)

    print(f"Generating dashboard for '{data['position']['title']}'…", flush=True)
    html = generate_html(data)

    with open(args.output, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"✓ Dashboard saved to: {args.output}")
    print(f"  {data['total']} candidates | {data['shortlisted']} shortlisted | "
          f"{data['on_hold']} on hold | {data['not_suitable']} not suitable")


if __name__ == "__main__":
    main()
