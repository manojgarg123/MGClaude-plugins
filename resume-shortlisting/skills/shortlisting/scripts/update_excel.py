#!/usr/bin/env python3
"""
update_excel.py — Create or update the candidate tracker Excel file.

Usage:
  python update_excel.py \
    --tracker "<path>" \
    --position-title "<role title>" \
    --name "<candidate name>" \
    --email "<email or N/A>" \
    --phone "<phone or N/A>" \
    --current-org "<current employer or N/A>" \
    --qualification "<highest degree e.g. M.Sc. Statistics>" \
    --years-experience "<e.g. 7>" \
    --resume "<resume filename>" \
    --skills-score <int 0-40> \
    --experience-score <int 0-30> \
    --education-score <int 0-20> \
    --fit-score <int 0-10> \
    --total-score <int 0-100> \
    --strengths "<text>" \
    --gaps "<text>" \
    --summary "<text>"
"""

import argparse
import os
from datetime import datetime


def get_or_create_workbook(tracker_path):
    try:
        import openpyxl
    except ImportError:
        os.system("pip install openpyxl --break-system-packages -q")
        import openpyxl

    HEADERS = [
        # ── Position (col 1) ───────────────────────────────
        "Position Title",
        # ── Candidate Details (cols 2-9) ───────────────────
        "Rank",
        "Candidate Name",
        "Email",
        "Phone",
        "Current Organisation",
        "Highest Qualification",
        "Years of Experience",
        "Resume File",
        # ── Fit Scores (cols 10-14) ────────────────────────
        "Skills Score (/40)",
        "Experience Score (/30)",
        "Education Score (/20)",
        "Overall Fit Score (/10)",
        "Total Score (/100)",
        # ── Assessment (cols 15-18) ────────────────────────
        "Key Strengths",
        "Key Gaps",
        "Summary",
        "Date Added",
    ]

    COL_WIDTHS = [
        25,                              # Position Title
        7, 22, 28, 16, 25, 28, 18, 25,  # Candidate Details
        18, 20, 18, 19, 15,             # Fit Scores
        35, 35, 55, 18,                 # Assessment
    ]

    # Column sets (1-based)
    POSITION_COLS = {1}
    DETAIL_COLS   = {2, 3, 4, 5, 6, 7, 8, 9}
    SCORE_COLS    = {10, 11, 12, 13, 14}
    TEXT_COLS     = {15, 16, 17}

    HEADER_COLORS = {
        "position": "4B0082",
        "detail":   "1F4E79",
        "score":    "1A5C2A",
        "text":     "7A5C00",
    }

    HEADER_FONT  = openpyxl.styles.Font(color="FFFFFF", bold=True, size=11)
    HEADER_ALIGN = openpyxl.styles.Alignment(
        horizontal="center", vertical="center", wrap_text=True)

    def header_fill(col_idx):
        if col_idx in POSITION_COLS:
            c = HEADER_COLORS["position"]
        elif col_idx in DETAIL_COLS:
            c = HEADER_COLORS["detail"]
        elif col_idx in SCORE_COLS:
            c = HEADER_COLORS["score"]
        else:
            c = HEADER_COLORS["text"]
        return openpyxl.styles.PatternFill(start_color=c, end_color=c, fill_type="solid")

    if os.path.exists(tracker_path):
        wb = openpyxl.load_workbook(tracker_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidate Tracker"
        ws.row_dimensions[1].height = 36

        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill(col_idx)
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN

        for col_idx, width in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"

    return wb, ws, HEADERS, POSITION_COLS, DETAIL_COLS, SCORE_COLS, TEXT_COLS


def find_existing_row(ws, candidate_name):
    """Return row number (1-based) if candidate already exists, else None.
    Candidate Name is column 3 (0-based index 2 in row tuple)."""
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell_val = row[2].value   # col 3 = Candidate Name
        if cell_val and str(cell_val).strip().lower() == candidate_name.strip().lower():
            return row[0].row
    return None


def add_or_update_candidate(ws, args, POSITION_COLS, DETAIL_COLS, SCORE_COLS, TEXT_COLS):
    import openpyxl

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    row_data = [
        # Position (col 1)
        args.position_title,
        # Candidate Details (cols 2-9)
        None,                       # Rank — filled after sort
        args.name,
        args.email,
        args.phone,
        args.current_org,
        args.qualification,
        args.years_experience,
        args.resume,
        # Fit Scores (cols 10-14)
        args.skills_score,
        args.experience_score,
        args.education_score,
        args.fit_score,
        args.total_score,
        # Assessment (cols 15-18)
        args.strengths,
        args.gaps,
        args.summary,
        timestamp,
    ]

    existing_row = find_existing_row(ws, args.name)
    updated = False

    if existing_row:
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=existing_row, column=col_idx, value=value)
        updated = True
        target_row = existing_row
    else:
        ws.append(row_data)
        target_row = ws.max_row

    # Apply row cell styling
    for col_idx in range(1, len(row_data) + 1):
        cell = ws.cell(row=target_row, column=col_idx)
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

        if col_idx in POSITION_COLS:
            cell.fill = openpyxl.styles.PatternFill(
                start_color="EDE7F6", end_color="EDE7F6", fill_type="solid")  # light lavender
        elif col_idx in SCORE_COLS:
            cell.fill = openpyxl.styles.PatternFill(
                start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # light green
        elif col_idx in TEXT_COLS:
            cell.fill = openpyxl.styles.PatternFill(
                start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # light yellow
        elif col_idx in DETAIL_COLS:
            cell.fill = openpyxl.styles.PatternFill(
                start_color="EAF4FB", end_color="EAF4FB", fill_type="solid")  # light blue

    # Colour-code the Total Score cell (col 14)
    total_cell = ws.cell(row=target_row, column=14)
    score = args.total_score
    if score >= 75:
        color = "00B050"
    elif score >= 50:
        color = "FF9900"
    else:
        color = "FF0000"
    total_cell.font = openpyxl.styles.Font(bold=True, color=color)

    return updated


def sort_and_rank(ws, total_col_idx=14, rank_col_idx=2):
    """Sort data rows by Total Score descending and refresh Rank column."""
    if ws.max_row < 2:
        return

    max_col = ws.max_column
    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            data_rows.append(list(row))

    data_rows.sort(
        key=lambda r: (float(r[total_col_idx - 1]) if r[total_col_idx - 1] is not None else 0),
        reverse=True,
    )

    for rank, row in enumerate(data_rows, start=1):
        row[rank_col_idx - 1] = rank

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, max_col + 1):
            ws.cell(row=row_idx, column=col_idx, value=None)

    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def main():
    parser = argparse.ArgumentParser(
        description="Create or update the candidate tracker Excel file")

    parser.add_argument("--tracker",        required=True, help="Path to tracker .xlsx")
    parser.add_argument("--position-title", default="N/A", help="Role title from JD")

    # Candidate details
    parser.add_argument("--name",             required=True)
    parser.add_argument("--email",            default="N/A")
    parser.add_argument("--phone",            default="N/A")
    parser.add_argument("--current-org",      default="N/A")
    parser.add_argument("--qualification",    default="N/A")
    parser.add_argument("--years-experience", default="N/A")
    parser.add_argument("--resume",           required=True)

    # Scores
    parser.add_argument("--skills-score",     type=int, required=True)
    parser.add_argument("--experience-score", type=int, required=True)
    parser.add_argument("--education-score",  type=int, required=True)
    parser.add_argument("--fit-score",        type=int, required=True)
    parser.add_argument("--total-score",      type=int, required=True)

    # Assessment
    parser.add_argument("--strengths", default="")
    parser.add_argument("--gaps",      default="")
    parser.add_argument("--summary",   default="")

    args = parser.parse_args()

    os.makedirs(os.path.dirname(os.path.abspath(args.tracker)), exist_ok=True)

    wb, ws, headers, POSITION_COLS, DETAIL_COLS, SCORE_COLS, TEXT_COLS = get_or_create_workbook(args.tracker)
    updated = add_or_update_candidate(ws, args, POSITION_COLS, DETAIL_COLS, SCORE_COLS, TEXT_COLS)
    sort_and_rank(ws)
    wb.save(args.tracker)

    action = "Updated existing" if updated else "Added new"
    print(f"✓ {action} candidate: {args.name} (Total Score: {args.total_score}/100)")
    print